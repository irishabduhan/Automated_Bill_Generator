import openpyxl
import os

def update_excel_file(file_path, updates, output_file_name):
    """
    Update the values in the INVOICE sheet of the given Excel file and save it with a new name.

    :param file_path: Path to the Excel file.
    :param updates: A list of tuples, where each tuple contains the cell address and the new value.
                    Example: [("B2", "New Name"), ("B5", "New PAN")]
    :param output_file_name: Name of the output Excel file.
    """
    # Load the workbook and the 'INVOICE' sheet
    workbook = openpyxl.load_workbook(file_path)
    if 'INVOICE' not in workbook.sheetnames:
        raise ValueError("Sheet 'INVOICE' does not exist in the workbook.")

    sheet = workbook['INVOICE']

    # Apply updates
    for cell_address, new_value in updates:
        sheet[cell_address] = new_value

    # Save the workbook with a new name
    output_file_path = os.path.join(os.path.dirname(file_path), output_file_name)
    workbook.save(output_file_path)
    print(f"Updates applied and saved to {output_file_path}")

def display_menu():
    """
    Display a menu for the user to select which fields to modify.
    """
    print("Select the fields to update:")
    print("1. Fill details from Excel file")
    print("2. Exit")

# Example usage
file_path = "C:\\Users\\risha\\Desktop\\Gourav Bhai\\Bill.xlsx"  # Path to your Excel file
source_file_path = "C:\\Users\\risha\\Desktop\\Gourav Bhai\\SourceData.xlsx"  # Path to the source data Excel file
updates = []

while True:
    display_menu()
    choice = input("Enter your choice: ")
    if choice == "1":
        try:
            # Load the source data workbook
            source_workbook = openpyxl.load_workbook(source_file_path)
            if 'Data' not in source_workbook.sheetnames:
                raise ValueError("Sheet 'Data' does not exist in the source workbook.")

            source_sheet = source_workbook['Data']

            # Iterate through rows in the source data sheet
            for row in source_sheet.iter_rows(min_row=2, values_only=True):
                # Assuming the columns are ordered as follows:
                # Name, PAN, Address Line 1, Address Line 2, Date, Description, Amount, Bill To, Company Address, Invoice No
                name, pan, address_line_1, address_line_2, date, description, amount, bill_to, company_address, invoice_no = row

                updates = [
                    ("B3", name),
                    ("B6", f"PAN - {pan}"),
                    ("B4", address_line_1),
                    ("B5", address_line_2),
                    ("H11", date),
                    ("B14", description),
                    ("H14", amount),
                    ("B8", f"Bill to - {bill_to}"),
                    ("B9", f"ADD - {company_address}"),
                    ("G11", invoice_no)
                ]

                # Generate output file name
                output_file_name = f"{name}_{pan}.xlsx"

                # Apply updates to the file
                update_excel_file(file_path, updates, output_file_name)

        except Exception as e:
            print(f"An error occurred: {e}")
    elif choice == "2":
        break
    else:
        print("Invalid choice. Please try again.")
